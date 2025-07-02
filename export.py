import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def export_to_excel(df: pd.DataFrame, filename: str):
    """
    Экспорт DataFrame с задачами в формате словарей в Excel с форматированием

    Параметры:
        df (pd.DataFrame): DataFrame с колонкой "Дата" и колонками сотрудников
        filename (str): Путь для сохранения файла (например, 'tasks.xlsx')
    """
    # Создаем копию DataFrame для обработки
    df_export = df.copy()

    # Конвертируем словари в строки для записи в Excel
    for col in df_export.columns[1:]:
        df_export[col] = df_export[col].apply(
            lambda x: "\n".join(f"{k}: {v}" for k, v in x.items()) if x else ""
        )

    # Создаем Excel writer
    writer = pd.ExcelWriter(filename, engine='openpyxl')
    df_export.to_excel(writer, index=False, sheet_name='Задачи')

    # Получаем объекты книги и листа
    workbook = writer.book
    worksheet = writer.sheets['Задачи']

    # Настраиваем стили
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Форматируем заголовки
    for col in range(1, len(df_export.columns) + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Настраиваем ширину столбцов
    for col_idx, column in enumerate(df_export.columns, 1):
        column_letter = get_column_letter(col_idx)

        # Автоподбор ширины с минимальным размером
        max_length = max(
            df_export[column].astype(str).apply(len).max(),
            len(column)
        )
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    # Форматируем ячейки с задачами
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                   min_col=2, max_col=worksheet.max_column):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    # Добавляем условное форматирование для статусов
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row,
                                   min_col=2, max_col=worksheet.max_column):
        for cell in row:
            if "2" in str(cell.value):
                cell.fill = green_fill
            elif "5" in str(cell.value):
                cell.fill = yellow_fill

    # Сохраняем файл
    writer.close()
    print(f"Файл успешно сохранен: {filename}")
